"""
TestcaseExporter 单元测试
===========================
覆盖: 文件生成 / 列验证 / 数据完整性 / 样式 / 边界情况
"""

import pytest
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from services.testcase_exporter import (
    TestcaseExporter,
    export_testcases,
    COLUMNS,
)


# ══════════════════════════════════════════════════════════
# Fixtures
# ══════════════════════════════════════════════════════════

@pytest.fixture
def sample_testcases() -> list[dict]:
    """标准测试用例数据。"""
    return [
        {
            "id": "TC-LOGIN-001",
            "title": "用户登录 - 正常密码 - 登录成功",
            "priority": "P0",
            "precondition": "测试账号 admin/Test@123 已注册，用户未登录",
            "steps": [
                "1. 打开登录页面 http://localhost/login",
                "2. 在用户名输入框输入 'admin'",
                "3. 在密码输入框输入 'Test@123'",
                "4. 点击「登录」按钮",
                "5. 验证页面跳转到系统首页",
            ],
            "expected_result": "登录成功，跳转到 /home，右上角显示用户名 'admin'",
        },
        {
            "id": "TC-LOGIN-002",
            "title": "用户登录 - 用户名为空 - 提示错误",
            "priority": "P1",
            "precondition": "打开登录页面",
            "steps": [
                "1. 打开登录页面",
                "2. 用户名输入框保持为空",
                "3. 在密码输入框输入 'Test@123'",
                "4. 点击「登录」按钮",
            ],
            "expected_result": "登录失败，提示'用户名不能为空'",
        },
        {
            "id": "TC-LOGIN-003",
            "title": "用户登录 - SQL注入 - 不泄露数据",
            "priority": "P0",
            "precondition": "打开登录页面",
            "steps": [
                "1. 打开登录页面",
                "2. 在用户名输入框输入 \"' OR '1'='1\"",
                "3. 在密码输入框输入 'test'",
                "4. 点击「登录」按钮",
                "5. 检查页面是否返回用户数据",
            ],
            "expected_result": "登录失败，不返回任何用户数据，提示'用户名或密码错误'",
        },
    ]


@pytest.fixture
def exporter() -> TestcaseExporter:
    """导出器实例（输出到临时目录）。"""
    import tempfile
    tmp = Path(tempfile.mkdtemp())
    return TestcaseExporter(output_dir=tmp)


# ══════════════════════════════════════════════════════════
# 基础功能测试
# ══════════════════════════════════════════════════════════

class TestExporterBasics:
    """基础功能测试。"""

    def test_export_creates_file(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """导出后文件存在且非空。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        assert filepath.exists()
        assert filepath.suffix == ".xlsx"
        assert filepath.stat().st_size > 0

    def test_export_filename_contains_module(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """文件名包含模块名称和时间戳。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        assert "用户登录" in filepath.name
        assert "测试用例" in filepath.name

    def test_export_empty_list_raises_error(self, exporter: TestcaseExporter) -> None:
        """空列表抛出 ValueError。"""
        with pytest.raises(ValueError) as exc_info:
            exporter.export([], "测试")
        assert "不能为空" in str(exc_info.value)

    def test_convenience_function(
        self, tmp_path: Path, sample_testcases: list
    ) -> None:
        """便捷函数 export_testcases 可正常使用。"""
        filepath = export_testcases(
            sample_testcases, "用户登录", output_dir=tmp_path
        )
        assert filepath.exists()
        assert filepath.suffix == ".xlsx"


# ══════════════════════════════════════════════════════════
# 内容校验测试
# ══════════════════════════════════════════════════════════

class TestExporterContent:
    """Excel 内容正确性测试。"""

    def _load(self, filepath: Path):
        """加载导出的工作簿。"""
        return load_workbook(str(filepath))

    def test_correct_columns(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """表头列名与定义一致。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        expected_headers = [c[0] for c in COLUMNS]
        for j, expected in enumerate(expected_headers, 1):
            actual = ws.cell(row=2, column=j).value
            assert actual == expected, f"列 {j}: 期望 '{expected}', 实际 '{actual}'"

    def test_row_count(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """数据行数 = 用例数量。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        # 第 1 行标题 + 第 2 行表头 + 数据行
        assert ws.max_row == 2 + len(sample_testcases)

    def test_data_integrity(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """导出后读取的数据与输入一致。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        # 检查第一条用例
        row = 3
        assert ws.cell(row=row, column=1).value == "TC-LOGIN-001"
        assert ws.cell(row=row, column=2).value == "用户登录"
        assert "正常密码" in str(ws.cell(row=row, column=3).value)
        assert ws.cell(row=row, column=4).value == "P0"
        assert "admin" in str(ws.cell(row=row, column=5).value)
        steps_val = str(ws.cell(row=row, column=6).value)
        assert "打开登录页面" in steps_val
        assert "admin" in steps_val
        assert "/home" in str(ws.cell(row=row, column=7).value)

    def test_priority_column_values(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """优先级列值为 P0/P1/P2/P3。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        priorities = set()
        for row in range(3, ws.max_row + 1):
            p = ws.cell(row=row, column=4).value
            priorities.add(p)
        assert priorities <= {"P0", "P1", "P2", "P3"}

    def test_module_column_filled(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """模块列（B 列）全部填充为 module_name。"""
        filepath = exporter.export(sample_testcases, "用户管理模块")
        wb = self._load(filepath)
        ws = wb.active

        for row in range(3, ws.max_row + 1):
            assert ws.cell(row=row, column=2).value == "用户管理模块"

    def test_steps_multiline_preserved(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """步骤列多行文本保留换行符。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        steps_text = str(ws.cell(row=3, column=6).value)
        assert "\n" in steps_text, "步骤应为多行文本"


# ══════════════════════════════════════════════════════════
# 样式校验测试
# ══════════════════════════════════════════════════════════

class TestExporterStyles:
    """Excel 样式测试。"""

    def _load(self, filepath: Path):
        return load_workbook(str(filepath))

    def test_freeze_panes(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """冻结窗格设置正确（A3 = 前两行固定）。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        assert ws.freeze_panes == "A3"

    def test_auto_filter_enabled(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """自动筛选已启用。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        assert ws.auto_filter.ref is not None
        assert "A2" in ws.auto_filter.ref

    def test_header_font_bold(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """表头字体加粗。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        for j in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=2, column=j)
            assert cell.font.bold is True, f"列 {j} 表头未加粗"

    def test_header_fill_color(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """表头背景为深蓝色。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        cell = ws.cell(row=2, column=1)
        assert cell.fill.start_color.rgb == "001F4E79"

    def test_body_alignment_wrap_text(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """数据行文本换行已启用。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        # 步骤列（F 列）和预期结果列（G 列）应启用换行
        for col in (6, 7):
            cell = ws.cell(row=3, column=col)
            assert cell.alignment.wrap_text is True, (
                f"列 {col} 未启用自动换行"
            )

    def test_priority_color_applied(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """P0 行优先级单元格有颜色标记。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        # 第 1 条用例 P0
        cell = ws.cell(row=3, column=4)
        assert cell.value == "P0"
        # 字体应为加粗
        assert cell.font.bold is True

    def test_column_width_auto_sized(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """列宽已设置（非默认值）。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        # A 列宽度应大于默认的 8
        width_a = ws.column_dimensions["A"].width
        assert width_a is not None and width_a > 8

    def test_title_row_merged(
        self, exporter: TestcaseExporter, sample_testcases: list
    ) -> None:
        """标题行合并了所有列。"""
        filepath = exporter.export(sample_testcases, "用户登录")
        wb = self._load(filepath)
        ws = wb.active

        assert len(ws.merged_cells.ranges) >= 1


# ══════════════════════════════════════════════════════════
# 边界情况测试
# ══════════════════════════════════════════════════════════

class TestExporterEdgeCases:
    """边界情况测试。"""

    def test_single_testcase(
        self, exporter: TestcaseExporter
    ) -> None:
        """只有 1 条用例也可正常导出。"""
        filepath = exporter.export([{
            "id": "TC-001",
            "title": "单条测试",
            "priority": "P2",
            "precondition": "无",
            "steps": ["1. 进行测试"],
            "expected_result": "测试完成",
        }], "测试模块")

        assert filepath.exists()
        wb = load_workbook(str(filepath))
        ws = wb.active
        assert ws.max_row == 3  # 标题 + 表头 + 1 数据

    def test_no_steps_field(self, exporter: TestcaseExporter) -> None:
        """没有 steps 字段时不出错。"""
        filepath = exporter.export([{
            "id": "TC-001",
            "title": "无步骤用例",
            "priority": "P2",
            "expected_result": "操作完成",
        }], "测试")

        assert filepath.exists()
        wb = load_workbook(str(filepath))
        ws = wb.active
        # steps 列应为空字符串
        steps_val = ws.cell(row=3, column=6).value
        assert steps_val is None or steps_val == ""

    def test_multiline_steps_row_height(
        self, exporter: TestcaseExporter
    ) -> None:
        """多步骤数据行高更大。"""
        filepath = exporter.export([{
            "id": "TC-MULTI-001",
            "title": "多步骤用例",
            "priority": "P1",
            "steps": [
                "1. 第一步", "2. 第二步", "3. 第三步",
                "4. 第四步", "5. 第五步", "6. 第六步",
                "7. 第七步", "8. 第八步",
            ],
            "expected_result": "全部步骤完成",
        }], "测试")

        wb = load_workbook(str(filepath))
        ws = wb.active
        # 行高应大于默认值
        assert ws.row_dimensions[3].height >= 20

    def test_special_chars_in_content(self, exporter: TestcaseExporter) -> None:
        """特殊字符正常导出（引号、尖括号、换行符）。"""
        filepath = exporter.export([{
            "id": "TC-SPECIAL-001",
            "title": "特殊字符 \" 测试 < > & ' ",
            "priority": "P1",
            "precondition": "数据包含 \"引号\" 和 <html> 标签",
            "steps": [
                "1. 输入 '<script>alert(1)</script>'",
                "2. 输入 \" OR '1'='1\"",
            ],
            "expected_result": "特殊字符被正确处理",
        }], "测试")

        wb = load_workbook(str(filepath))
        ws = wb.active
        steps = str(ws.cell(row=3, column=6).value)
        assert "script" in steps
        assert "alert" in steps
        assert "OR" in steps
        # SQL 注入特征字符应保留
        assert any(c in steps for c in ("1=1", "1'='1", "1＝1"))
