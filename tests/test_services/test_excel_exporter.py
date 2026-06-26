from openpyxl import load_workbook

from services.excel_exporter import ExcelExporter, ExportData


def test_functional_testcase_sheet_uses_business_columns(tmp_path) -> None:
    filepath = ExcelExporter(tmp_path).export(ExportData(
        project_name="demo",
        features=[],
        testpoints=[],
        testcases=[{
            "case_id": "TC-001",
            "title": "登录 - 正确账号密码 - 登录成功",
            "testpoint_description": "用户登录",
            "priority": "P0",
            "precondition": "已打开登录页",
            "steps": ["1. 输入账号", "2. 点击登录"],
            "body": "admin/123456",
            "expected": "进入首页",
        }],
        testcase_mode="functional",
    ))

    ws = load_workbook(filepath)["测试用例清单"]
    headers = [ws.cell(row=2, column=i).value for i in range(1, 9)]

    assert headers == [
        "用例编号", "用例标题", "模块/项目", "优先级",
        "前置条件", "测试步骤", "测试数据", "预期结果",
    ]
    assert ws.cell(row=3, column=7).value == "admin/123456"


def test_api_testcase_sheet_keeps_interface_columns(tmp_path) -> None:
    filepath = ExcelExporter(tmp_path).export(ExportData(
        project_name="demo",
        features=[],
        testpoints=[],
        testcases=[{
            "case_id": "TC-API-001",
            "title": "登录接口 - 正确账号密码 - 返回 token",
            "priority": "P0",
            "precondition": "服务已启动",
            "method": "POST",
            "url": "/api/login",
            "headers": '{"Content-Type": "application/json"}',
            "body": '{"username": "admin"}',
            "expected": "HTTP 200",
        }],
        testcase_mode="api",
    ))

    ws = load_workbook(filepath)["测试用例清单"]
    headers = [ws.cell(row=2, column=i).value for i in range(1, 11)]

    assert headers == [
        "用例编号", "用例标题", "模块/项目", "优先级", "前置条件",
        "请求方法", "URL", "请求头", "请求体", "预期结果",
    ]
    assert ws.cell(row=3, column=6).value == "POST"
