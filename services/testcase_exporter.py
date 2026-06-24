"""
测试用例 Excel 导出服务
=========================
将测试用例列表导出为标准 Excel 模板（7 列）。

特性：
- 自动列宽（根据内容自适应）
- 自动换行（多行文本完整显示）
- 冻结首行（滚动时表头保持可见）
- 样式美化（深蓝表头、斑马纹、优先级颜色标记）
"""

import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import OUTPUT_DIR

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════
# 样式常量
# ══════════════════════════════════════════════════════════

# ── 颜色 ──────────────────────────────────────────
HEADER_BG = "1F4E79"       # 深蓝背景
HEADER_FG = "FFFFFF"       # 白色文字
ROW_EVEN = "F2F7FB"        # 偶数行浅蓝
ROW_ODD = "FFFFFF"         # 奇数行白色
BORDER_COLOR = "C0C0C0"    # 边框灰色

# ── 优先级颜色 ────────────────────────────────────
PRIORITY_STYLES = {
    "P0": {"bg": "FFC7CE", "font": "9C0006"},  # 红色背景 + 深红文字
    "P1": {"bg": "FFEB9C", "font": "9C6500"},  # 黄色背景 + 深黄文字
    "P2": {"bg": "C6EFCE", "font": "006100"},  # 绿色背景 + 深绿文字
    "P3": {"bg": "E2EFDA", "font": "3F3F3F"},  # 浅绿背景 + 灰色文字
}

# ── 字体 ──────────────────────────────────────────
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color=HEADER_FG)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
BODY_FONT = Font(name="微软雅黑", size=10)
BODY_FONT_BOLD = Font(name="微软雅黑", size=10, bold=True)

# ── 填充 ──────────────────────────────────────────
HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
TITLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ROW_EVEN_FILL = PatternFill(start_color=ROW_EVEN, end_color=ROW_EVEN, fill_type="solid")

# ── 对齐 ──────────────────────────────────────────
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)

# ── 边框 ──────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

# ── 列定义 ────────────────────────────────────────
COLUMNS = [
    ("用例序号", 18),  # A: TC-LOGIN-001
    ("用例标题", 36),  # B: 被测对象-场景-预期行为
    ("模块/项目", 16), # C: 所属模块
    ("优先级", 8),     # D: P0/P1/P2/P3
    ("前置条件", 30),  # E: 前置条件
    ("测试步骤", 52),  # F: 测试步骤（多行编号）
    ("测试数据", 25),  # G: 建议测试数据
    ("预期结果", 40),  # H: 预期结果
]


# ══════════════════════════════════════════════════════════
# TestcaseExporter
# ══════════════════════════════════════════════════════════

class TestcaseExporter:
    """测试用例 Excel 导出器。

    生成包含单个工作表的标准测试用例模板。

    Usage:
        exporter = TestcaseExporter()
        filepath = exporter.export(
            testcases=[...],
            module_name="用户管理",
        )
        print(filepath)  # outputs/用户管理_20250101_120000.xlsx
    """

    COL_ID = 1        # A: 用例序号
    COL_TITLE = 2     # B: 用例标题
    COL_MODULE = 3    # C: 模块/项目
    COL_PRIORITY = 4  # D: 优先级
    COL_PRECOND = 5   # E: 前置条件
    COL_STEPS = 6     # F: 测试步骤
    COL_DATA = 7      # G: 测试数据
    COL_EXPECTED = 8  # H: 预期结果

    def __init__(self, output_dir: Path | None = None) -> None:
        """初始化导出器。

        Args:
            output_dir: 输出目录（None 则使用配置中的默认值）
        """
        self._output_dir = output_dir or OUTPUT_DIR
        self._output_dir.mkdir(parents=True, exist_ok=True)

    def export(
        self,
        testcases: list[dict[str, Any]],
        module_name: str = "通用",
    ) -> Path:
        """导出测试用例为 Excel 文件。

        Args:
            testcases: 测试用例列表，每项含 id/title/precondition/steps/expected_result/priority
            module_name: 模块名称（用于文件名和 B 列填充）

        Returns:
            生成的 Excel 文件路径

        Raises:
            ValueError: testcases 为空
        """
        if not testcases:
            raise ValueError("测试用例列表不能为空")

        logger.info(
            "开始导出 Excel: 模块=%s, 用例数=%d", module_name, len(testcases)
        )

        wb = Workbook()

        # ── 创建工作表 ──────────────────────────────
        ws = wb.active
        ws.title = "测试用例"

        # ── 标题行 ──────────────────────────────────
        self._write_title_row(ws, module_name)

        # ── 表头行 ──────────────────────────────────
        self._write_header_row(ws)

        # ── 数据行 ──────────────────────────────────
        self._write_data_rows(ws, testcases, module_name)

        # ── 后处理 ──────────────────────────────────
        self._apply_styles(ws)

        # ── 保存 ────────────────────────────────────
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(
            c if c.isalnum() or c in "._- " else "_" for c in module_name
        )
        filename = f"测试用例_{safe_name}_{timestamp}.xlsx"
        filepath = self._output_dir / filename

        wb.save(str(filepath))
        logger.info("Excel 导出完成: %s", filepath)

        return filepath

    # ── 写入方法 ────────────────────────────────────

    def _write_title_row(self, ws: Worksheet, module_name: str) -> None:
        """写入大标题行（第 1 行，合并单元格居中）。

        Args:
            ws: 工作表
            module_name: 模块名称
        """
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(COLUMNS),
        )
        cell = ws.cell(
            row=1, column=1,
            value=f"测试用例 — {module_name}",
        )
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = TITLE_FILL
        ws.row_dimensions[1].height = 32

    def _write_header_row(self, ws: Worksheet) -> None:
        """写入表头行（第 2 行，冻结行）。

        Args:
            ws: 工作表
        """
        for j, (header, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=2, column=j, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        ws.row_dimensions[2].height = 28

    def _write_data_rows(
        self,
        ws: Worksheet,
        testcases: list[dict[str, Any]],
        module_name: str,
    ) -> None:
        """写入数据行（从第 3 行开始）。

        Args:
            ws: 工作表
            testcases: 测试用例列表
            module_name: 模块名称（填充到 B 列）
        """
        for i, tc in enumerate(testcases):
            row = i + 3  # 数据从第 3 行开始
            is_even = i % 2 == 1

            # ── 处理步骤文本 ────────────────────────
            steps = tc.get("steps", [])
            if isinstance(steps, list):
                steps_text = "\n".join(
                    s if s.strip().startswith(tuple(f"{n}." for n in range(20)))
                    else f"{idx}. {s}"
                    for idx, s in enumerate(steps, 1)
                )
            else:
                steps_text = str(steps) if steps else ""

            # ── 写入各列（8列标准格式）──────────────
            values = [
                tc.get("case_id", tc.get("id", "")),                  # A: 用例序号
                tc.get("title", ""),                                  # B: 用例标题
                tc.get("module", module_name),                        # C: 模块/项目
                tc.get("priority", "P2"),                             # D: 优先级
                tc.get("precondition", "无"),                         # E: 前置条件
                steps_text,                                           # F: 测试步骤
                tc.get("test_data", ""),                              # G: 测试数据
                tc.get("expected_result", tc.get("expected", "")),   # H: 预期结果
            ]

            for j, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER

                # ── 对齐方式 ────────────────────────
                if j in (1, 4):  # 序号列 + 优先级列居中
                    cell.alignment = CENTER_ALIGN
                elif j in (6, 7, 8):  # 步骤/数据/预期换行
                    cell.alignment = BODY_ALIGN
                else:
                    cell.alignment = BODY_ALIGN

                # ── 偶数行背景色 ────────────────────
                if is_even:
                    cell.fill = ROW_EVEN_FILL

                # ── 优先级颜色标记 ──────────────────
                if j == 4:  # D 列
                    priority = str(val).strip()
                    if priority in PRIORITY_STYLES:
                        style = PRIORITY_STYLES[priority]
                        cell.fill = PatternFill(
                            start_color=style["bg"],
                            end_color=style["bg"],
                            fill_type="solid",
                        )
                        cell.font = Font(
                            name="微软雅黑", size=10, bold=True,
                            color=style["font"],
                        )

            # ── 行高：根据步骤数量自适应 ────────────
            step_count = len(steps) if isinstance(steps, list) else 1
            ws.row_dimensions[row].height = max(20, 15 * max(step_count, 2))

    def _apply_styles(self, ws: Worksheet) -> None:
        """应用全局样式：列宽、冻结窗格、自动筛选。

        Args:
            ws: 工作表
        """
        # ── 自动列宽 ──────────────────────────────
        for j, (_, default_width) in enumerate(COLUMNS, 1):
            # 计算该列实际最大宽度
            max_width = default_width
            for row in ws.iter_rows(
                min_row=3, max_col=j, max_row=ws.max_row, values_only=True
            ):
                cell_val = str(row[0]) if row[0] else ""
                # 取每行最长的一行（多行文本）
                for line in cell_val.split("\n"):
                    # 中文字符算 2 个宽度
                    char_width = sum(2 if ord(c) > 127 else 1 for c in line)
                    max_width = max(max_width, char_width + 4)

            # 限制最大宽度避免过宽
            max_width = min(max_width, 80)
            ws.column_dimensions[get_column_letter(j)].width = max_width

        # ── 冻结首行（第 3 行开始滚动时前 2 行固定）───
        ws.freeze_panes = "A3"

        # ── 自动筛选 ──────────────────────────────
        last_col = get_column_letter(len(COLUMNS))
        ws.auto_filter.ref = f"A2:{last_col}{ws.max_row}"


# ══════════════════════════════════════════════════════════
# 便捷函数
# ══════════════════════════════════════════════════════════

def export_testcases(
    testcases: list[dict[str, Any]],
    module_name: str = "通用",
    output_dir: Path | None = None,
) -> Path:
    """导出测试用例为 Excel 的便捷函数。

    Args:
        testcases: 测试用例列表
        module_name: 模块名称
        output_dir: 输出目录

    Returns:
        生成的 Excel 文件路径

    Example:
        filepath = export_testcases([
            {"id": "TC-001", "title": "正常登录", "steps": ["1. 输入", "2. 点击"], ...},
        ], "用户登录")
    """
    exporter = TestcaseExporter(output_dir=output_dir)
    return exporter.export(testcases, module_name)
