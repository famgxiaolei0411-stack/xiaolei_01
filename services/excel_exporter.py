"""
Excel 导出器 — 将功能点、测试点、测试用例导出为 Excel
==========================================================
生成格式化的 .xlsx 文件，包含多工作表、样式、冻结窗格。
"""

import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from typing import Any
from config import OUTPUT_DIR

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════
# 样式常量
# ══════════════════════════════════════════════════════════

# ── 颜色定义 ──────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
SUB_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUB_HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")

BODY_FONT = Font(name="微软雅黑", size=10)
BODY_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── 优先级颜色映射 ────────────────────────────────
PRIORITY_FILLS = {
    "P0": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # 红色
    "P1": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # 黄色
    "P2": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # 绿色
    "P3": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # 浅绿
}
PRIORITY_FONTS = {
    "P0": Font(name="微软雅黑", size=10, bold=True, color="9C0006"),
    "P1": Font(name="微软雅黑", size=10, bold=True, color="9C6500"),
    "P2": Font(name="微软雅黑", size=10, color="006100"),
    "P3": Font(name="微软雅黑", size=10, color="3F3F3F"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


# ══════════════════════════════════════════════════════════
# 导出器
# ══════════════════════════════════════════════════════════

@dataclass
class ExportData:
    """导出数据结构（统一使用 dict，不再依赖 V1 dataclass）。

    Attributes:
        project_name: 项目名称
        features: 功能点列表 [{"module","name","description","priority","preconditions","business_rules"},...]
        testpoints: 测试点列表 [{"feature_name","category","description","expected_result","test_data","priority"},...]
        testcases: 测试用例列表 [{"case_id","title","precondition","steps","expected","priority","case_type"},...]
    """
    project_name: str
    features: list[dict[str, Any]]
    testpoints: list[dict[str, Any]]
    testcases: list[dict[str, Any]]


class ExcelExporter:
    """Excel 导出器。

    生成 3 个工作表:
    1. 功能点清单 — 模块、名称、描述、优先级、前置条件
    2. 测试点清单 — 关联功能点、测试类型、描述、预期结果
    3. 测试用例清单 — 用例编号、标题、步骤、预期、优先级、类型

    使用方式:
        exporter = ExcelExporter()
        filepath = exporter.export(data)
    """

    def __init__(self, output_dir: Path | None = None) -> None:
        """初始化导出器。

        Args:
            output_dir: 输出目录（None 则使用默认配置）
        """
        self._output_dir = output_dir or OUTPUT_DIR
        self._output_dir.mkdir(parents=True, exist_ok=True)

    def export(self, data: ExportData) -> Path:
        """导出为 Excel 文件。

        Args:
            data: 导出数据

        Returns:
            生成的 Excel 文件路径
        """
        logger.info("开始导出 Excel: 项目=%s", data.project_name)

        wb = Workbook()

        # ── 创建工作表 ──────────────────────────────
        self._build_features_sheet(wb, data.features)
        self._build_testpoints_sheet(wb, data.testpoints)
        self._build_testcases_sheet(wb, data.testcases)

        # 删除默认空工作表
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # 激活第一个工作表
        wb.active = 0

        # ── 保存文件 ────────────────────────────────
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(
            c if c.isalnum() or c in "._- " else "_"
            for c in data.project_name
        )
        filename = f"{safe_name}_{timestamp}.xlsx"
        filepath = self._output_dir / filename

        wb.save(str(filepath))
        logger.info("Excel 导出完成: %s", filepath)

        return filepath

    # ── 工作表构建 ──────────────────────────────────

    def _build_features_sheet(
        self,
        wb: Workbook,
        features: list[dict[str, Any]],
    ) -> None:
        """构建「功能点清单」工作表。"""
        ws = wb.active
        ws.title = "功能点清单"

        headers = ["序号", "所属模块", "功能点名称", "功能描述", "优先级", "前置条件", "业务规则"]
        col_widths = [6, 16, 20, 50, 8, 30, 30]

        self._write_header(ws, headers, col_widths, "需求功能点清单")

        for i, feat in enumerate(features, 1):
            row = i + 2  # 数据从第3行开始
            values = [
                i,
                feat.get("module", ""),
                feat.get("name", ""),
                feat.get("description", ""),
                feat.get("priority", "P2"),
                "\n".join(feat.get("preconditions", [])) if feat.get("preconditions") else "-",
                "\n".join(feat.get("business_rules", [])) if feat.get("business_rules") else "-",
            ]
            for j, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=val)
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGNMENT
                cell.border = THIN_BORDER

                # 优先级列特殊样式
                if j == 5:
                    cell.alignment = CENTER_ALIGNMENT
                    p = str(val).strip()
                    if p in PRIORITY_FILLS:
                        cell.fill = PRIORITY_FILLS[p]
                        cell.font = PRIORITY_FONTS[p]

        self._finalize_sheet(ws, len(features) + 2, len(headers))

    def _build_testpoints_sheet(
        self,
        wb: Workbook,
        testpoints: list[dict[str, Any]],
    ) -> None:
        """构建「测试点清单」工作表。"""
        ws = wb.create_sheet("测试点清单")

        headers = ["序号", "关联功能点", "测试类型", "测试点描述", "预期结果", "建议测试数据", "优先级"]
        col_widths = [6, 20, 14, 50, 40, 25, 8]

        self._write_header(ws, headers, col_widths, "测试点清单")

        for i, tp in enumerate(testpoints, 1):
            row = i + 2
            values = [
                i,
                tp.get("feature_name", ""),
                tp.get("category", ""),
                tp.get("description", ""),
                tp.get("expected_result", ""),
                tp.get("test_data") or "-",
                tp.get("priority", "P1"),
            ]
            for j, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=val)
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGNMENT
                cell.border = THIN_BORDER

                if j == 7:  # 优先级
                    cell.alignment = CENTER_ALIGNMENT
                    p = str(val).strip()
                    if p in PRIORITY_FILLS:
                        cell.fill = PRIORITY_FILLS[p]
                        cell.font = PRIORITY_FONTS[p]

        self._finalize_sheet(ws, len(testpoints) + 2, len(headers))

    def _build_testcases_sheet(
        self,
        wb: Workbook,
        testcases: list[dict[str, Any]],
    ) -> None:
        """构建「测试用例清单」工作表 — 接口测试 10 列格式。

        用例编号 | 用例标题 | 模块/项目 | 优先级 | 前置条件 |
        请求方法 | URL | 请求头 | 请求体 | 预期结果
        """
        ws = wb.create_sheet("测试用例清单")

        headers = [
            "用例编号", "用例标题", "模块/项目",
            "优先级", "前置条件", "请求方法",
            "URL", "请求头", "请求体", "预期结果",
        ]
        col_widths = [18, 36, 16, 8, 22, 10, 30, 22, 40, 40]

        self._write_header(ws, headers, col_widths, "接口测试用例清单")

        for i, tc in enumerate(testcases, 1):
            row = i + 2
            case_id = tc.get("case_id", "") or f"TC-{i:03d}"

            values = [
                case_id,                                    # A: 用例编号
                tc.get("title", ""),                        # B: 用例标题
                tc.get("testpoint_description") or "-",     # C: 模块/项目
                tc.get("priority", "P1"),                   # D: 优先级
                tc.get("precondition") or "无",              # E: 前置条件
                tc.get("method") or "GET",                  # F: 请求方法
                tc.get("url") or "-",                       # G: URL
                tc.get("headers") or "-",                   # H: 请求头
                tc.get("body") or "-",                      # I: 请求体
                tc.get("expected", ""),                     # J: 预期结果
            ]
            for j, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=j, value=val)
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGNMENT
                cell.border = THIN_BORDER

                if j == 1:  # 编号
                    cell.font = Font(name="微软雅黑", size=10, bold=True)
                    cell.alignment = CENTER_ALIGNMENT
                if j == 4:  # 优先级
                    cell.alignment = CENTER_ALIGNMENT
                    p = str(val).strip()
                    if p in PRIORITY_FILLS:
                        cell.fill = PRIORITY_FILLS[p]
                        cell.font = PRIORITY_FONTS[p]
                if j == 6:  # 请求方法
                    cell.alignment = CENTER_ALIGNMENT

        self._finalize_sheet(ws, len(testcases) + 2, len(headers))

    # ── 工具方法 ────────────────────────────────────

    def _write_header(
        self,
        ws: Worksheet,
        headers: list[str],
        col_widths: list[int],
        title: str,
    ) -> None:
        """写入标题行和表头。

        Args:
            ws: 工作表
            headers: 列标题
            col_widths: 列宽
            title: 表格标题
        """
        # ── 大标题行 ────────────────────────────────
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(headers),
        )
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.row_dimensions[1].height = 30

        # ── 表头行 ──────────────────────────────────
        for j, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=2, column=j, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(j)].width = width

        ws.row_dimensions[2].height = 25

    def _finalize_sheet(
        self,
        ws: Worksheet,
        total_rows: int,
        total_cols: int,
    ) -> None:
        """工作表最终设置：冻结窗格、自动筛选、行高。

        Args:
            ws: 工作表
            total_rows: 总行数
            total_cols: 总列数
        """
        # ── 冻结窗格（标题+表头）─────────────────────
        ws.freeze_panes = "A3"

        # ── 自动筛选 ────────────────────────────────
        ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{total_rows}"

        # ── 数据行高 ────────────────────────────────
        for row in range(3, total_rows + 1):
            ws.row_dimensions[row].height = max(25, 15 * 3)  # 适应换行文本


# ══════════════════════════════════════════════════════════
# 便捷函数
# ══════════════════════════════════════════════════════════

def export_to_excel(
    project_name: str,
    features: list[dict[str, Any]],
    testpoints: list[dict[str, Any]],
    testcases: list[dict[str, Any]],
) -> Path:
    """导出 Excel 的便捷函数。

    Args:
        project_name: 项目名称
        features: 功能点列表
        testpoints: 测试点列表
        testcases: 测试用例列表

    Returns:
        生成的 Excel 文件路径
    """
    data = ExportData(
        project_name=project_name,
        features=features,
        testpoints=testpoints,
        testcases=testcases,
    )
    exporter = ExcelExporter()
    return exporter.export(data)
